import os
import subprocess
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from rest_framework import viewsets
from rest_framework.views import APIView
from django.core.files import File
from django.core.files.storage import default_storage
from datetime import timedelta
from django.db import IntegrityError
from .models import Employee ,DoctorVideo, Doctor,EmployeeLoginHistory, VideoTemplates,DoctorOutputVideo
from .serializers import EmployeeLoginSerializer,EmployeeSerializer,DoctorSerializer,DoctorVideoSerializer, VideoTemplatesSerializer,DoctorOutputVideoSerializer
import pandas as pd # type: ignore
from django.conf import settings
from django.core.files import File
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from rest_framework_simplejwt.views import TokenRefreshView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.response import Response
from rest_framework import status
import requests
import uuid
from distutils.util import strtobool
from rest_framework.permissions import IsAuthenticated
from .serializers import DoctorVideoSerializer
from datetime import datetime, date
import random
import string
import openpyxl
from django.utils import timezone
from django.http import HttpResponse
import logging
from rest_framework.pagination import PageNumberPagination
logger = logging.getLogger(__name__)





class Pagination_class(PageNumberPagination):
    page_size = 10  # You can change this default page size
    page_size_query_param = 'page_size'
    max_page_size = 100

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer



def get_tokens_for_employee(employee):
    refresh = RefreshToken.for_user(employee)
    
    # Set expiration time for access token
    refresh.access_token.set_exp(lifetime=timedelta(hours=1))  # Expiry in 1 hour

    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'access_token_exp': refresh.access_token.payload['exp'],  # Expiry time of access token
    }

@api_view(['POST'])
def employee_login_api(request):
    serializer = EmployeeLoginSerializer(data=request.data)
    if serializer.is_valid():
        employee_id = serializer.validated_data['employee_id']
        try:
            employee = Employee.objects.get(employee_id=employee_id)

            if not employee.status:
                return Response({
                    'status': 'error',
                    'message': 'Your account is inactive. Please contact the admin department.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            employee.login_date = timezone.now()
            employee.has_logged_in = True
            employee.save(update_fields=['login_date', 'has_logged_in']) 
            
            
            tokens = get_tokens_for_employee(employee)

            EmployeeLoginHistory.objects.create(
                employee=employee,
                employee_identifier=employee.employee_id,
                name=f"{employee.first_name} {employee.last_name}",
                email=employee.email,
                phone=employee.phone,
                department=employee.department,
                user_type=employee.user_type
            )


            return Response({
                'status': 'success',
                'message': 'Login successful',
                'tokens': tokens,
                'employee': {
                    'id':employee.id,
                    'employee_id': employee.employee_id,
                    'name': f"{employee.first_name} {employee.last_name}",
                    'email': employee.email,
                    'department': employee.department,
                    'user_type': employee.user_type
                }
            }, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Invalid employee ID'
            }, status=status.HTTP_401_UNAUTHORIZED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])  # Important for file uploads
def add_doctor(request):
    serializer = DoctorSerializer(data=request.data)
    print(serializer,"serializer")
    if serializer.is_valid():
        serializer.save()
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'status': 'error', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


###### Video creation #######

# class VideoGenViewSet(viewsets.ModelViewSet):
#     queryset = DoctorVideo.objects.all()
#     serializer_class = DoctorVideoSerializer

#     def perform_create(self, serializer):
#         request = self.request
#         doctor = serializer.save()

#         # Skip if doctor has no image
#         if not doctor.image:
#             return

#         # Use template_id from request if given
#         template_id = request.data.get("template_id")
#         template = None

#         if template_id:
#             template = VideoTemplates.objects.filter(id=template_id).first()
#         else:
#             template = VideoTemplates.objects.filter(status=True).first()

#         if not template:
#             return 

#         # Prepare output path
#         output_filename = f"{doctor.id}_{template.id}_output.mp4"
#         output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
#         os.makedirs(output_dir, exist_ok=True)
#         output_path = os.path.join(output_dir, output_filename)

#         try:
#             # Call the same logic from GenerateDoctorOutputVideoView
#             GenerateDoctorOutputVideoView().generate_custom_video(
#                 main_video_path=template.template_video.path,
#                 image_path=doctor.image.path,
#                 name=doctor.name,
#                 clinic=doctor.clinic,
#                 city=doctor.city,
#                 specialization_key=doctor.specialization_key,
#                 time_duration=template.time_duration,
#                 state=doctor.state,
#                 output_path=output_path,
#                 resolution=template.resolution,
#                 base_x=template.base_x_axis,
#                 base_y=template.base_y_axis,
#                 line_spacing=template.line_spacing,
#                 overlay_x=template.overlay_x,
#                 overlay_y=template.overlay_y,
#             )

#             # Save generated video
#             relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)

#             DoctorOutputVideo.objects.create(
#                 doctor=doctor,
#                 template=template,
#                 video_file=relative_path
#             )
#         except Exception as e:
#             print(f"Error generating video after doctor creation: {e}")


class VideoGenViewSet(viewsets.ModelViewSet):
    queryset = DoctorVideo.objects.all()
    serializer_class = DoctorVideoSerializer

    def parse_time_duration(self, time_duration_str):
        """
        Parse time duration string like "2-6,65-70" into slots list [(2, 6), (65, 70)]
        """
        if not time_duration_str or not time_duration_str.strip():
            raise ValueError("Time duration cannot be empty")
        
        try:
            slots = []
            time_ranges = time_duration_str.strip().split(',')
            
            for time_range in time_ranges:
                start_str, end_str = time_range.strip().split('-')
                start_time = int(start_str.strip())
                end_time = int(end_str.strip())
                
                if start_time >= end_time:
                    raise ValueError(f"Start time ({start_time}) must be less than end time ({end_time})")
                
                if start_time < 0:
                    raise ValueError(f"Start time ({start_time}) cannot be negative")
                
                slots.append((start_time, end_time))
            
            return slots
            
        except ValueError as e:
            if "not enough values to unpack" in str(e) or "too many values to unpack" in str(e):
                raise ValueError(f"Invalid time duration format: '{time_duration_str}'. Use format like '10-15' or '10-15,46-50'")
            raise e
        except Exception as e:
            raise ValueError(f"Error parsing time duration '{time_duration_str}': {str(e)}")


    def perform_create(self, serializer):
        # print("bedbug: perform_create called")
        logger.info("VideoGenViewSet perform_create called")
        request = self.request
        doctor = serializer.save()
        # print("bedbug: Doctor saved:", doctor)
        logger.info(f"Doctor saved: {doctor.name} (ID: {doctor.id})")

        # Skip if doctor has no image
        if not doctor.image:
            # print("bedbug: Doctor has no image, skipping video generation")
            logger.warning(f"Doctor {doctor.name} has no image, skipping video generation")
            return

        # Use template_id from request if given
        template_id = request.data.get("template_id")
        # print("bedbug: template_id from request:", template_id)
        logger.info(f"Template ID from request: {template_id}")
        template = None

        if template_id:
            # print("bedbug: Fetching template by template_id")
            logger.info("Fetching template by provided template_id")
            template = VideoTemplates.objects.filter(id=template_id).first()
        else:
            # print("bedbug: Fetching first active template")
            logger.info("No template_id provided, fetching first active template")
            template = VideoTemplates.objects.filter(status=True).first()

        if not template:
            # print("bedbug: No template found, aborting video generation")
            logger.error("No template found, aborting video generation")
            return 

        # Prepare output path
        output_filename = f"{doctor.id}_{template.id}_output.mp4"
        # print("bedbug: output_filename:", output_filename)
        logger.info(f"Output filename: {output_filename}")
        output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
        # print("bedbug: output_dir:", output_dir)
        logger.info(f"Output directory: {output_dir}")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)
        # print("bedbug: output_path:", output_path)
        logger.info(f"Full output path: {output_path}")

        try:
            # print("bedbug: Calling generate_custom_video...")
            logger.info("Calling self.generate_custom_video...")
            self.generate_custom_video(
                main_video_path=template.template_video.path,
                image_path=doctor.image.path,
                name=doctor.name,
                clinic=doctor.clinic,
                city=doctor.city,
                specialization_key=doctor.specialization_key,
                time_duration=template.time_duration,
                state=doctor.state,
                output_path=output_path,
                resolution=template.resolution,
                base_x=template.base_x_axis,
                base_y=template.base_y_axis,
                line_spacing=template.line_spacing,
                overlay_x=template.overlay_x,
                overlay_y=template.overlay_y,
            )
            # print("bedbug: Video generated successfully")
            logger.info("Video generated successfully")
            # Save generated video
            relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)
            # print("bedbug: relative_path:", relative_path)
            logger.info(f"Relative path for database: {relative_path}")
            DoctorOutputVideo.objects.create(
                doctor=doctor,
                template=template,
                video_file=relative_path
            )
            # print("bedbug: DoctorOutputVideo record created")
            logger.info("DoctorOutputVideo database record created successfully")
        except Exception as e:
            # print(f"bedbug: Error generating video after doctor creation: {e}")
            logger.error(f"VideoGenViewSet error generating video: {e}")
            logger.error(f"Exception type: {type(e).__name__}")

    def generate_custom_video(self, main_video_path, image_path, name, clinic, city, specialization_key, state, output_path,
                      time_duration="5-10,45-50", resolution="415x410", base_x="(main_w/2)-160", base_y="(main_h/2)-60", line_spacing="60", overlay_x="350",
                      overlay_y="70"):
        logger.info("=== VideoGenViewSet.generate_custom_video STARTED ===")
        logger.info(f"Parameters - doctor: {name}, template: {time_duration}, resolution: {resolution}")
        # logger.info(f"Starting video generation for doctor: {name}")

        # Resource validation  
        if not os.path.exists(main_video_path):
            logger.error(f"Template video not found: {main_video_path}")
            raise Exception(f"Template video not found: {main_video_path}")

        if not os.path.exists(image_path):
            logger.error(f"Doctor image not found: {image_path}")
            raise Exception(f"Doctor image not found: {image_path}")
        # print("bedbug: Entered generate_custom_video")
        logger.info(f"Starting video generation for doctor: {name}")
        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        # print("bedbug: temp_dir created:", temp_dir)
        logger.info(f"Temp directory created: {temp_dir}")
        fps = 30
        fade_duration = 3
        # print("bedbug: Parsing time_duration for slots...")
        logger.info(f"Parsing time duration: {time_duration}")
        slots = self.parse_time_duration(time_duration)
        # print("bedbug: slots:", slots)
        logger.info(f"Parsed slots: {slots}")
        temp_videos = []
        for i, (start, end) in enumerate(slots):
            # print(f"bedbug: Creating temp video for slot {i}: start={start}, end={end}")
            logger.info(f"Creating temp video {i} for slot: {start}-{end}s")
            duration = end - start
            total_frames = duration * fps
            # print(f"bedbug: duration={duration}, total_frames={total_frames}")
            logger.debug(f"Slot {i} duration: {duration}s, frames: {total_frames}")
            zoom_effect = f"zoompan=z='1+0.00003*in':x='(iw/2)-(iw/zoom/2)':y='(ih/2)-(ih/zoom/2)':d={total_frames}:s={resolution}:fps={fps}"
            # fade_effect = f"format=rgba,fade=t=in:st=0:d={fade_duration}:alpha=1,fade=t=out:st={duration-fade_duration}:d={fade_duration}:alpha=1"
            # byvaso
            fade_out_start = max(0, duration - fade_duration)
            fade_effect = (
                f"format=rgba,"
                f"fade=t=in:st=0:d={fade_duration}:alpha=1,"
                f"fade=t=out:st={fade_out_start}:d={fade_duration}:alpha=1"
            )

            vf = f"scale={resolution},{zoom_effect},{fade_effect}"
            temp_video = os.path.join(temp_dir, f"temp_image_vid_{i}.mp4")
            # print("bedbug: Running ffmpeg for temp video:", temp_video)
            logger.debug(f"Running FFmpeg for temp video: {temp_video}")
            try:
                result = subprocess.run([
                    "ffmpeg", "-loop", "1", "-i", image_path,
                    "-vf", vf, "-t", str(duration), "-y", temp_video
                ], check=True, capture_output=True, text=True)
            except subprocess.CalledProcessError as e:
                logger.error(f"Failed to create temp video {i}: {e.stderr}")
                raise Exception(f"Failed to create temp video: {e.stderr}")
            temp_videos.append((temp_video, start, end))
        # print("bedbug: temp_videos created:", temp_videos)
        logger.info(f"Created {len(temp_videos)} temp videos")
        text_lines = [name, specialization_key, clinic, city, state]
        # font = "RobotoSlab-Medium.ttf"
# Try to find font file in different locations
        font_paths = [
            os.path.join(settings.BASE_DIR, "fonts", "RobotoSlab-Medium.ttf"),
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Common Linux font
            "/System/Library/Fonts/Arial.ttf",  # macOS fallback
            "arial.ttf"  # Windows fallback
        ]

        font = None
        for font_path in font_paths:
            if os.path.exists(font_path):
                font = font_path
                break

        if not font:
            logger.warning("No font file found, using system default")
            # Use system default font or remove font specification
            font = "arial"  # Let ffmpeg find system font
        # print("bedbug: text_lines:", text_lines)
        logger.info(f"Text lines: {text_lines}")
        # print(f"bedbug: Looking for font at: {font}")
        font_status = os.path.exists(font) if '/' in font else 'system font'
        logger.info(f"Using font: {font} (exists: {font_status})")

        # print(f"bedbug: Font exists: {os.path.exists(font) if '/' in font else 'system font'}")
        logger.info(f"Template video: {main_video_path} (exists: {os.path.exists(main_video_path)})")

        # print(f"bedbug: Template video path: {main_video_path}")
        # print(f"bedbug: Template video exists: {os.path.exists(main_video_path)}")
        logger.info(f"Template video exists: {os.path.exists(main_video_path)}")
        text_filters = []
        for start, end in slots:
            alpha_expr = f"if(lt(t\\,{start}+3),(t-{start})/3,if(lt(t\\,{end}-3),1,({end}-t)/3))"
            for j, text in enumerate(text_lines):
                offset = 132
                y_pos = f"(main_h - ({len(text_lines)}*{line_spacing}) + {j}*{line_spacing} - {offset})"
                x_pos = "(main_w/2)-300"
                # drawtext = (
                #     f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                #     f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                # )
                if font:
                    drawtext = (
                        f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                        f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                    )
                else:
                    drawtext = (
                        f"drawtext=text='{text}':fontcolor=black:fontsize=40:"
                        f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                    )
                # print(f"bedbug: drawtext filter for line {j}:", drawtext)
                logger.debug(f"Text filter line {j}: {drawtext}")
                text_filters.append(drawtext)
        overlay_x1 = f"(main_w-overlay_w)/2-{int(overlay_x)}"
        overlay_y1 = f"(main_h-overlay_h)/2+{int(overlay_y)}"
        # print("bedbug: overlay_x1:", overlay_x1, "overlay_y1:", overlay_y1)
        logger.debug(f"Overlay positions - x1: {overlay_x1}, y1: {overlay_y1}")


        #! tODAYS WORKS AFTER DEPLOYEMNT ISSUES 
        # filter_complex = (
        #     f"[0:v][1:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[0][0]},{slots[0][1]})'[v1];"
        #     f"[v1][2:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[1][0]},{slots[1][1]})'[v2];"
        #     f"[v2]{','.join(text_filters)}[v]"
        # )

        # Build overlay filters dynamically based on number of slots
        overlay_filters = []
        for i, (start, end) in enumerate(slots):
            if i == 0:
                input_label = "[0:v]"
                output_label = f"[v{i+1}]"
            else:
                input_label = f"[v{i}]"
                output_label = f"[v{i+1}]"
            
            overlay_filter = f"{input_label}[{i+1}:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{start},{end})'{output_label}"
            overlay_filters.append(overlay_filter)

        # Final filter combines overlays with text
        final_input = f"[v{len(slots)}]"
        filter_complex = f"{';'.join(overlay_filters)};{final_input}{','.join(text_filters)}[v]"
        # print("bedbug: filter_complex:", filter_complex)
        logger.debug(f"Filter complex: {filter_complex}")
        # cmd = [
        #     "ffmpeg", "-i", main_video_path,
        #     "-i", temp_videos[0][0],
        #     "-i", temp_videos[1][0],
        #     "-filter_complex", filter_complex,
        #     "-map", "[v]",
        #     "-map", "0:a?",
        #     "-c:v", "libx264", "-c:a", "copy", "-y", output_path
        # ]
        cmd = ["ffmpeg", "-i", main_video_path]
        # Add all temp video inputs dynamically
        for temp_video, _, _ in temp_videos:
            cmd.extend(["-i", temp_video])

        cmd.extend([
            "-filter_complex", filter_complex,
            "-map", "[v]",
            "-map", "0:a?",
            "-c:v", "libx264", "-c:a", "copy", "-y", output_path
        ])
        # print("bedbug: Running final ffmpeg command:", cmd)
        logger.info("Running final FFmpeg command")
        logger.debug(f"FFmpeg command: {' '.join(cmd)}")
        # subprocess.run(cmd, check=True)
        try:
            result = subprocess.run(cmd, check=True, capture_output=True, text=True)
        except subprocess.CalledProcessError as e:
            logger.error(f"Final FFmpeg failed: {e.stderr}")
            raise Exception(f"Video generation failed: {e.stderr}")
        for temp_video, _, _ in temp_videos:
            # print("bedbug: Removing temp video:", temp_video)
            logger.debug(f"Removing temp video: {temp_video}")
            os.remove(temp_video)
        logger.info("=== VideoGenViewSet.generate_custom_video COMPLETED ===")

    

class DoctorVideoViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DoctorVideo.objects.all().order_by('-created_at')
    serializer_class = DoctorVideoSerializer
    pagination_class = Pagination_class



@api_view(['POST'])
@parser_classes([MultiPartParser])
def bulk_upload_employees(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(excel_file)
        required_columns = {'first_name', 'last_name', 'email', 'phone', 'department', 'date_joined'}

        if not required_columns.issubset(df.columns):
            return Response({'error': f'Missing required columns: {required_columns - set(df.columns)}'}, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        def generate_employee_id(first_name, last_name):
            base = f"EMP{first_name[0].upper()}{last_name[0].upper()}"
            while True:
                suffix = ''.join(random.choices(string.digits, k=4))
                emp_id = f"{base}{suffix}"
                if not Employee.objects.filter(employee_id=emp_id).exists():
                    return emp_id

        for index, row in df.iterrows():
            row_number = index + 2

            first_name = str(row.get('first_name', '')).strip()
            last_name = str(row.get('last_name', '')).strip()
            email = str(row.get('email', '')).strip()
            phone = str(row.get('phone', '')).strip()

            if not first_name or not last_name or not email:
                errors.append({'row': row_number, 'error': 'Missing first name, last name, email or phone'})
                skipped += 1
                continue

            employee_id = str(row.get('employee_id')).strip()

            data = {
                'employee_id': employee_id,
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'phone': phone,
                'department': row.get('department', '') or 'Unknown',
                'date_joined': row.get('date_joined') if not pd.isna(row.get('date_joined')) else datetime.today().date()
            }

            serializer = EmployeeSerializer(data=data)
            if serializer.is_valid():
                try:
                    serializer.save()
                    created += 1
                except IntegrityError:
                    errors.append({'row': row_number, 'error': f'Duplicate email or other unique constraint'})
                    skipped += 1
            else:
                errors.append({'row': row_number, 'error': serializer.errors})
                skipped += 1

        return Response({
            'created': created,
            'skipped': skipped,
            'errors': errors
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


   
class DoctorListByEmployee(APIView):

    permission_classes = [IsAuthenticated]
      
    def get(self, request):
        # Get the employee_id from query parameters
        employee_id = request.GET.get('employee_id')

        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=400)

        try:
            # Find the employee by employee_id
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found for this employee_id."}, status=404)

        # Filter doctors associated with the employee
        doctors = Doctor.objects.filter(employee=employee)

        # Serialize the data
        serializer = DoctorSerializer(doctors, many=True)
        return Response(serializer.data)
    

class DoctorVideoListView(APIView):

    def get(self, request):
        employee_id = request.GET.get('employee_id')

        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({"detail": "Employee not found for this employee_id."}, status=404)

        doctors = DoctorVideo.objects.filter(employee=employee).order_by('-created_at')  # sort by newest first

        # Apply pagination
        paginator = Pagination_class()
        page = paginator.paginate_queryset(doctors, request)

        serializer = DoctorVideoSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

class DoctorVideoGeneration(APIView):

    def post(self, request):
        # Use request.GET.get('id') in production — hardcoded for testing here
        doctor_id = request.data.get('id')

        if not doctor_id:
            return Response({"detail": "doctor_id is required."}, status=400)

        try:
            # Fetch a single DoctorVideo instance
            doctor_data = DoctorVideo.objects.get(id=doctor_id)
            print(doctor_data, "doctor_data-------------")
            # Call your video generation logic
            generate_video_for_doctor(doctor_data)

            return Response({
                "detail": "Video generation successful.",
                "video_path": request.build_absolute_uri(doctor_data.output_video.url) if doctor_data.output_video else None
            })

        except DoctorVideo.DoesNotExist:
            return Response({"detail": "Doctor not found for this doctor_id."}, status=404)

        except Exception as e:
            return Response({"detail": f"Error during video generation: {str(e)}"}, status=500)


class CustomTokenRefreshView(TokenRefreshView):
    def post(self, request, *args, **kwargs):
        refresh_token = request.data.get('refresh', None)
        if not refresh_token:
            return Response({
                'status': 'error',
                'message': 'Refresh token is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Verify the refresh token and get new access token
            refresh = RefreshToken(refresh_token)

            print(refresh)

            new_refresh_token = str(refresh)

            new_access_token = str(refresh.access_token)

            return Response({
                'status': 'success',
                'message': 'Token refreshed successfully',
                'access': new_access_token,
                'refresh': new_refresh_token,
                'access_token_exp': refresh.access_token.payload['exp'],  # Expiry time of access token
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'status': 'error',
                'message': 'Invalid refresh token'
            }, status=status.HTTP_401_UNAUTHORIZED)
        



def generate_video_for_doctor(doctor):
    print(f"Generating video for doctor {doctor.name}")

    try:
        if not doctor.image:
            print(f"No image for doctor {doctor.name}, skipping video.")
            return

        main_video_path = os.path.join(settings.MEDIA_ROOT, "Health111.mp4")
        image_path = doctor.image.path
        print(image_path,"image_path")
        output_dir = os.path.join(settings.MEDIA_ROOT, "output")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{doctor.id}_output.mp4")

        # Generate video using custom logic
        VideoGenViewSet().generate_custom_video(
            main_video_path,
            image_path,
            doctor.name,
            doctor.clinic,
            doctor.city,
            doctor.specialization,
            doctor.state,
            output_path
        )

        # Save the output video as a FileField to the Doctor model (or however you store it)
        with open(output_path, 'rb') as f:


            doctor.output_video.save(f"{doctor.id}_output.mp4", File(f), save=True)

        
        BASE_URL = "https://api.videomaker.digielvestech.in/"
        
        doctor.output_video_url = f"{BASE_URL}{doctor.output_video.url}"

        print(doctor.output_video_url,"doctor.output_video_url")
        doctor.save()

        print(f"Video generated and saved for doctor {doctor.name}")

    except Exception as e:
        print(f"Error generating video for doctor {doctor.name}: {e}")




@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_doctors(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(excel_file)
        required_columns = {'name', 'clinic', 'city', 'specialization', 'state'}

        if not required_columns.issubset(df.columns):
            return Response({'error': f'Missing required columns: {required_columns - set(df.columns)}'},
                            status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        for index, row in df.iterrows():
            row_number = index + 2
            name = str(row.get('name', '')).strip()
            clinic = str(row.get('clinic', '')).strip()
            city = str(row.get('city', '')).strip()
            specialization = str(row.get('specialization', '')).strip()
            state = str(row.get('state', '')).strip()

            # Skip empty rows
            if not name or not clinic or not city or not specialization or not state:
                skipped += 1
                errors.append({'row': row_number, 'error': 'Required fields are missing'})
                continue


            designation = str(row.get('designation', '')).strip()
            mobile_number = str(row.get('mobile_number', '')).strip()
            whatsapp_number = str(row.get('whatsapp_number', '')).strip()
            description = str(row.get('description', '')).strip()
            image_path = str(row.get('image_url', '')).strip()
            employee = row.get('emp_id')

            image_file = None
            if image_path and os.path.exists(image_path):
                image_file = File(open(image_path, 'rb'), name=os.path.basename(image_path))

            doctor_data = {
                'name': name,
                'clinic': clinic,
                'city': city,
                'specialization': specialization,
                'state': state,
                'image': image_file,
                'designation': designation,
                'mobile_number': mobile_number,
                'whatsapp_number': whatsapp_number,
                'description': description,
                'employee': employee,
            }

            serializer = DoctorSerializer(data=doctor_data)
            if serializer.is_valid():
                try:
                    doctor = serializer.save()
                    generate_video_for_doctor(doctor)  # ✅ One video per doctor
                    created += 1
                except IntegrityError:
                    skipped += 1
                    errors.append({'row': row_number, 'error': 'Integrity error during save'})
            else:
                skipped += 1
                errors.append({'row': row_number, 'error': serializer.errors})

        return Response({
            'created': created,
            'skipped': skipped,
            'errors': errors
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


class DoctorVideoExportExcelView(APIView):

    def get(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Doctor Videos"

        headers = [
            "Name", "Designation", "Clinic", "City", "State", "Image URL",
            "Specialization", "Mobile Number", "WhatsApp Number", "Description","Template Name",
            "Output Video URL", "Created At", "Employee ID", "Employee Name", "RBM Name"
        ]
        sheet.append(headers)

        doctor_videos = DoctorVideo.objects.all()

        for video in doctor_videos:
            # Get latest DoctorOutputVideo instance
            latest_output = video.doctoroutputvideo_set.order_by('-created_at').first()

            if latest_output and latest_output.video_file:
                output_video_url = request.build_absolute_uri(latest_output.video_file.url)
                template_name = latest_output.template.name if latest_output.template else ""
            elif video.output_video:
                output_video_url = request.build_absolute_uri(video.output_video.url)
                template_name = ""
            else:
                output_video_url = ""
                template_name = ""

            rbm_name = (
                f"{video.employee.rbm.first_name} {video.employee.rbm.last_name}"
                if video.employee and video.employee.rbm else ""
            )

            sheet.append([
                video.name,
                video.designation,
                video.clinic,
                video.city,
                video.state,
                request.build_absolute_uri(video.image.url) if video.image else "",
                video.specialization,
                video.mobile_number,
                video.whatsapp_number,
                video.description,
                template_name,
                output_video_url,  # ✅ Now coming from DoctorOutputVideo
                video.created_at.strftime('%Y-%m-%d %H:%M:%S') if video.created_at else "",
                video.employee.employee_id if video.employee else "",
                f"{video.employee.first_name} {video.employee.last_name}" if video.employee else "",
                rbm_name,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=doctor_videos.xlsx'
        workbook.save(response)
        return response



class EmployeeExportExcelView(APIView):

    def get(self, request):
        # Create Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Employees"

        # Add headers
        headers = [
            "Employee ID", "First Name", "Last Name", "Email", "Phone",
            "Department", "Date Joined", "User Type", "Status"
        ]
        sheet.append(headers)

        # Populate data
        for emp in Employee.objects.all():
            sheet.append([
                emp.employee_id,
                emp.first_name,
                emp.last_name or "",
                emp.email or "",
                emp.phone or "",
                emp.department or "",
                emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else "",
                emp.user_type,
                "Active" if emp.status else "Inactive"
            ])

        # Prepare Excel response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
        workbook.save(response)
        return response


@api_view(['GET'])
def total_employee_count(request):
    count = Employee.objects.count()
    return Response({
        "total_employees": count
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def todays_active_employees(request):
    today = timezone.now().date()

    # Get all employees who have login history with login_time today
    employees = Employee.objects.filter(
        login_history__login_time__date=today
    ).distinct()

    # Count of those employees
    count = employees.count()

    data = [{
        'employee_id': emp.employee_id,
        'name': f"{emp.first_name} {emp.last_name}",
        'email': emp.email,
        'department': emp.department,
        'user_type': emp.user_type,
    } for emp in employees]

    return Response({
        'date': str(today),
        'active_employee_count': count,
        'active_employees': data
    })


class TodaysActiveEmployeeExcelExport(APIView):
    def get(self, request):
        today = timezone.now().date()

        # Get employees who logged in today
        employees = Employee.objects.filter(
            login_history__login_time__date=today
        ).distinct()

        # Create Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Today's Active Employees"

        # Add headers
        headers = [
            "Employee ID", "First Name", "Last Name", "Email", "Phone",
            "Department", "Date Joined", "User Type", "Status"
        ]
        sheet.append(headers)

        # Populate data
        for emp in employees:
            sheet.append([
                emp.employee_id,
                emp.first_name,
                emp.last_name or "",
                emp.email or "",
                emp.phone or "",
                emp.department or "",
                emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else "",
                emp.user_type,
                "Active" if emp.status else "Inactive"
            ])

        # Prepare response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"active_employees_{today}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response
    




@api_view(['GET'])
def doctors_with_output_video_count(request):
    count = DoctorVideo.objects.filter(
        Q(output_video__isnull=False) & ~Q(output_video='')
    ).count()
    return Response({
        "doctor_with_output_video_count": count
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def doctors_with_output_video_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Doctor Videos"

    # Updated header row (added "RBM Name")
    headers = [
        "Name", "Designation", "Clinic", "City", "State", "Image URL",
        "Specialization", "Mobile Number", "WhatsApp Number", "Description",
        "Output Video URL", "Created At", "Employee ID", "Employee Name", "RBM Name"
    ]
    sheet.append(headers)

    # Get doctor videos with output video only
    # doctor_videos = DoctorVideo.objects.filter(
    #     Q(output_video__isnull=False) & ~Q(output_video='')
    # )
    doctor_videos = DoctorVideo.objects.all()

    for video in doctor_videos:
        # Safely get RBM name
        rbm_name = (
            f"{video.employee.rbm.first_name} {video.employee.rbm.last_name}"
            if video.employee and video.employee.rbm else ""
        )

        sheet.append([
            video.name,
            video.designation,
            video.clinic,
            video.city,
            video.state,
            request.build_absolute_uri(video.image.url) if video.image else "",
            video.specialization,
            video.mobile_number,
            video.whatsapp_number,
            video.description,
            request.build_absolute_uri(video.output_video.url) if video.output_video else "",
            video.created_at.strftime('%Y-%m-%d %H:%M:%S') if video.created_at else "",
            video.employee.employee_id if video.employee else "",
            f"{video.employee.first_name} {video.employee.last_name}" if video.employee else "",
            rbm_name,
        ])

    # Return Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=doctor_videos.xlsx'
    workbook.save(response)
    return response


@api_view(['GET'])
def doctors_count(request):
    count = DoctorVideo.objects.filter().count()
    return Response({
        "doctor_count": count
    }, status=status.HTTP_200_OK)



class VideoTemplateAPIView(APIView):

    def get(self, request, pk=None):
        if pk:
            template = get_object_or_404(VideoTemplates, pk=pk)
            serializer = VideoTemplatesSerializer(template)
        else:
            status_param = request.query_params.get('status')

            templates = VideoTemplates.objects.all()

            if status_param is not None:
                try:
                    status_bool = bool(strtobool(status_param))  # Converts "true"/"false" to True/False
                    templates = templates.filter(status=status_bool)
                except ValueError:
                    return Response({"error": "Invalid status value. Use true or false."}, status=status.HTTP_400_BAD_REQUEST)

            templates = templates.order_by('-created_at')
            serializer = VideoTemplatesSerializer(templates, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = VideoTemplatesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk)
        serializer = VideoTemplatesSerializer(template, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            print(serializer.data)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        template = get_object_or_404(VideoTemplates, pk=pk)
        template.delete()
        return Response({"detail": "Deleted successfully."}, status=status.HTTP_204_NO_CONTENT)



# class GenerateDoctorOutputVideoView(APIView):
#     def parse_time_duration(self, time_duration_str):
#         """
#         Parse time duration string like "2-6,65-70" into slots list [(2, 6), (65, 70)]
#         """
#         if not time_duration_str or not time_duration_str.strip():
#             raise ValueError("Time duration cannot be empty")
        
#         try:
#             slots = []
#             # Split by comma to get individual time ranges
#             time_ranges = time_duration_str.strip().split(',')
            
#             for time_range in time_ranges:
#                 # Split by dash to get start and end times
#                 start_str, end_str = time_range.strip().split('-')
#                 start_time = int(start_str.strip())
#                 end_time = int(end_str.strip())
                
#                 # Validate times
#                 if start_time >= end_time:
#                     raise ValueError(f"Start time ({start_time}) must be less than end time ({end_time})")
                
#                 if start_time < 0:
#                     raise ValueError(f"Start time ({start_time}) cannot be negative")
                
#                 slots.append((start_time, end_time))
            
#             return slots
            
#         except ValueError as e:
#             if "not enough values to unpack" in str(e) or "too many values to unpack" in str(e):
#                 raise ValueError(f"Invalid time duration format: '{time_duration_str}'. Use format like '10-15' or '10-15,46-50'")
#             raise e
#         except Exception as e:
#             raise ValueError(f"Error parsing time duration '{time_duration_str}': {str(e)}")

#     def post(self, request):
#         doctor_id = request.data.get("doctor_id")
#         template_id = request.data.get("template_id")

#         if not doctor_id:
#             return Response({"error": "doctor_id is required."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             doctor = DoctorVideo.objects.get(id=doctor_id)
#         except DoctorVideo.DoesNotExist:
#             return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

#         # Use default template if not provided
#         if template_id:
#             try:
#                 template = VideoTemplates.objects.get(id=template_id)
#             except VideoTemplates.DoesNotExist:
#                 return Response({"error": "Template not found."}, status=status.HTTP_404_NOT_FOUND)
#         else:
#             template = VideoTemplates.objects.filter(status=True).first()
#             if not template:
#                 return Response({"error": "No default template available."}, status=status.HTTP_400_BAD_REQUEST)

#         if not doctor.image:
#             return Response({"error": "Doctor does not have an image."}, status=status.HTTP_400_BAD_REQUEST)

#         image_path = doctor.image.path

#         # Prepare output path
#         random_key = uuid.uuid4().hex[:8]
#         output_filename = f"{doctor.id}_{template.id}_{random_key}_output.mp4"

#         output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
#         os.makedirs(output_dir, exist_ok=True)
#         output_path = os.path.join(output_dir, output_filename)

#         try:
#             self.generate_custom_video(
#             main_video_path=template.template_video.path,
#             image_path=image_path,
#             name=doctor.name,
#             clinic=doctor.clinic,
#             city=doctor.city,
#             specialization_key=doctor.specialization_key,
#             state=doctor.state,
#             output_path=output_path,
#             time_duration=template.time_duration,
#             resolution=template.resolution,
#             base_x=template.base_x_axis,
#             base_y=template.base_y_axis,
#             line_spacing=template.line_spacing,
#             overlay_x=template.overlay_x,
#             overlay_y=template.overlay_y,
#         )

#             relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)

#             output_video = DoctorOutputVideo.objects.create(
#                 doctor=doctor,
#                 template=template,
#                 video_file=relative_path
#             )

#             # with open(output_path, 'rb') as f:
#             #     output_video = DoctorOutputVideo.objects.create(
#             #         doctor=doctor,
#             #         template=template,
#             #         video_file=File(f, name=output_filename)
#             #     )

#             serializer = DoctorOutputVideoSerializer(output_video)
#             return Response(serializer.data, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             print(f"Video generation failed: {e}")
#             return Response({"error": "Video generation failed.", "details": str(e)},
#                             status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#     def generate_custom_video(self, main_video_path, image_path, name, clinic, city, specialization_key, state, output_path,
#                       time_duration, resolution="415x410", base_x="(main_w/2)-160", base_y="(main_h/2)-60", line_spacing="60",overlay_x="350",
#                       overlay_y="70"):
#         temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
#         os.makedirs(temp_dir, exist_ok=True)

#         fps = 30
#         fade_duration = 3   
#         slots = self.parse_time_duration(time_duration)

#         temp_videos = []
#         for i, (start, end) in enumerate(slots):
#             duration = end - start
#             total_frames = duration * fps

#             zoom_effect = f"zoompan=z='1+0.00003*in':x='(iw/2)-(iw/zoom/2)':y='(ih/2)-(ih/zoom/2)':d={total_frames}:s={resolution}:fps={fps}"
#             fade_effect = f"format=rgba,fade=t=in:st=0:d={fade_duration}:alpha=1,fade=t=out:st={duration-fade_duration}:d={fade_duration}:alpha=1"
#             vf = f"scale={resolution},{zoom_effect},{fade_effect}"

#             temp_video = os.path.join(temp_dir, f"temp_image_vid_{i}.mp4")
#             subprocess.run([
#                 "ffmpeg", "-loop", "1", "-i", image_path,
#                 "-vf", vf, "-t", str(duration), "-y", temp_video
#             ], check=True)

#             temp_videos.append((temp_video, start, end))

#         # === Drawtext settings ===
#         text_lines = [name, specialization_key, clinic, city, state]
#         font = "RobotoSlab-Medium.ttf"

#         text_filters = []
#         for start, end in slots:
#             alpha_expr = f"if(lt(t\\,{start}+3),(t-{start})/3,if(lt(t\\,{end}-3),1,({end}-t)/3))"
#             for j, text in enumerate(text_lines):
#                 # y_pos = f"{base_y}+{j}*{line_spacing}"
#                 # drawtext = (
#                 #     f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
#                 #     f"x={base_x}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
#                 # )



#                 offset = 132  # adjust this value to move text further up or down
#                 y_pos = f"(main_h - ({len(text_lines)}*{line_spacing}) + {j}*{line_spacing} - {offset})"

#                 x_pos= "(main_w/2)-300"
               
#                 drawtext = (
#                     f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
#                     f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
#                 )

#                 text_filters.append(drawtext)

#         overlay_x1 = f"(main_w-overlay_w)/2-{int(overlay_x)}"
#         overlay_y1 = f"(main_h-overlay_h)/2+{int(overlay_y)}"

#         filter_complex = (
#             f"[0:v][1:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[0][0]},{slots[0][1]})'[v1];"
#             f"[v1][2:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[1][0]},{slots[1][1]})'[v2];"
#             f"[v2]{','.join(text_filters)}[v]"
#         )

#         cmd = [
#             "ffmpeg", "-i", main_video_path,
#             "-i", temp_videos[0][0],
#             "-i", temp_videos[1][0],
#             "-filter_complex", filter_complex,
#             "-map", "[v]",
#             "-map", "0:a?",
#             "-c:v", "libx264", "-c:a", "copy", "-y", output_path
#         ]
#         subprocess.run(cmd, check=True)

#         for temp_video, _, _ in temp_videos:
#             os.remove(temp_video)

#     def get(self, request):
#         doctor_id = request.query_params.get("doctor_id")
#         employee_id = request.query_params.get("employee_id")

#         videos = DoctorOutputVideo.objects.all().order_by('-id')  # Latest created first

#         if doctor_id:
#             videos = videos.filter(doctor_id=doctor_id)

#         if employee_id:
#             videos = videos.filter(doctor__employee_id=employee_id)

#         paginator = Pagination_class()
#         paginated_videos = paginator.paginate_queryset(videos, request)

#         serializer = DoctorOutputVideoSerializer(paginated_videos, many=True)
#         return paginator.get_paginated_response(serializer.data)

class GenerateDoctorOutputVideoView(APIView):
    def parse_time_duration(self, time_duration_str):
        # print("bedbug: Entered parse_time_duration with:", time_duration_str)
        logger.info(f"Parsing time duration: {time_duration_str}")
        if not time_duration_str or not time_duration_str.strip():
            # print("bedbug: Empty time_duration_str, raising ValueError")
            logger.error("Empty time_duration_str, raising ValueError")
            raise ValueError("Time duration cannot be empty")
        try:
            slots = []
            time_ranges = time_duration_str.strip().split(',')
            # print("bedbug: time_ranges:", time_ranges)
            logger.debug(f"Time ranges: {time_ranges}")
            for time_range in time_ranges:
                start_str, end_str = time_range.strip().split('-')
                start_time = int(start_str.strip())
                end_time = int(end_str.strip())
                # print(f"bedbug: time_range={time_range}, start_time={start_time}, end_time={end_time}")
                logger.debug(f"Processing range {time_range}: start={start_time}, end={end_time}")
                if start_time >= end_time:
                    # print("bedbug: Start time >= end time, raising ValueError")
                    logger.error("Start time >= end time, raising ValueError")
                    raise ValueError(f"Start time ({start_time}) must be less than end time ({end_time})")
                if start_time < 0:
                    # print("bedbug: Start time < 0, raising ValueError")
                    logger.error("Start time < 0, raising ValueError")
                    raise ValueError(f"Start time ({start_time}) cannot be negative")
                slots.append((start_time, end_time))
            # print("bedbug: Returning slots:", slots)
            logger.info(f"Parsed slots: {slots}")

            return slots
        except ValueError as e:
            # print("bedbug: ValueError in parse_time_duration:", e)
            logger.error(f"ValueError in parse_time_duration: {e}")
            if "not enough values to unpack" in str(e) or "too many values to unpack" in str(e):
                raise ValueError(f"Invalid time duration format: '{time_duration_str}'. Use format like '10-15' or '10-15,46-50'")
            raise e
        except Exception as e:
            # print("bedbug: General exception in parse_time_duration:", e)
            logger.error(f"General exception in parse_time_duration: {e}")
            raise ValueError(f"Error parsing time duration '{time_duration_str}': {str(e)}")

    def post(self, request):
        # print("bedbug: POST called, request.data:", request.data)
        logger.info(f"POST request received with data: {request.data}")
        doctor_id = request.data.get("doctor_id")
        template_id = request.data.get("template_id")
        # print(f"bedbug: doctor_id={doctor_id}, template_id={template_id}")
        logger.info(f"Processing doctor_id={doctor_id}, template_id={template_id}")
        if not doctor_id:
            # print("bedbug: doctor_id is missing, returning error")
            logger.error("doctor_id is missing, returning error")
            return Response({"error": "doctor_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            # print("bedbug: Fetching DoctorVideo object...")
            logger.info("Fetching DoctorVideo object...")
            doctor = DoctorVideo.objects.get(id=doctor_id)
            # print("bedbug: DoctorVideo found:", doctor)
            logger.info(f"DoctorVideo found: {doctor}")
        except DoctorVideo.DoesNotExist:
            # print("bedbug: DoctorVideo does not exist")
            logger.error("DoctorVideo does not exist")
            return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

        if template_id:
            try:
                # print("bedbug: Fetching VideoTemplates with provided template_id...")
                logger.info("Fetching VideoTemplates with provided template_id...")
                template = VideoTemplates.objects.get(id=template_id)
                # print("bedbug: Template found:", template)
                logger.info(f"Template found: {template}")
            except VideoTemplates.DoesNotExist:
                # print("bedbug: VideoTemplates does not exist")
                logger.error("VideoTemplates does not exist")
                return Response({"error": "Template not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            # print("bedbug: Fetching first default active VideoTemplate...")
            logger.info("Fetching first default active VideoTemplate...")
            template = VideoTemplates.objects.filter(status=True).first()
            if not template:
                # print("bedbug: No default template available")
                logger.error("No default template available")
                return Response({"error": "No default template available."}, status=status.HTTP_400_BAD_REQUEST)

        if not doctor.image:
            # print("bedbug: Doctor has no image")
            logger.error("Doctor has no image")
            return Response({"error": "Doctor does not have an image."}, status=status.HTTP_400_BAD_REQUEST)

        image_path = doctor.image.path
        # print("bedbug: doctor.image.path:", image_path)
        logger.info(f"Doctor image path: {image_path}")
        random_key = uuid.uuid4().hex[:8]
        output_filename = f"{doctor.id}_{template.id}_{random_key}_output.mp4"
        # print("bedbug: output_filename:", output_filename)
        logger.info(f"Output filename: {output_filename}")
        output_dir = os.path.join(settings.MEDIA_ROOT, "output", str(doctor.employee.id), str(doctor.id))
        # print("bedbug: output_dir:", output_dir)
        logger.info(f"Output directory: {output_dir}")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)
        # print("bedbug: output_path:", output_path)
        logger.info(f"Output path: {output_path}")

        try:
            # print("bedbug: Calling generate_custom_video...")
            logger.info("Starting video generation...")
            self.generate_custom_video(
                main_video_path=template.template_video.path,
                image_path=image_path,
                name=doctor.name,
                clinic=doctor.clinic,
                city=doctor.city,
                specialization_key=doctor.specialization_key,
                state=doctor.state,
                output_path=output_path,
                time_duration=template.time_duration,
                resolution=template.resolution,
                base_x=template.base_x_axis,
                base_y=template.base_y_axis,
                line_spacing=template.line_spacing,
                overlay_x=template.overlay_x,
                overlay_y=template.overlay_y,
            )
            relative_path = os.path.relpath(output_path, settings.MEDIA_ROOT)
            # print("bedbug: relative_path:", relative_path)
            logger.info(f"Relative path: {relative_path}")
            output_video = DoctorOutputVideo.objects.create(
                doctor=doctor,
                template=template,
                video_file=relative_path
            )
            # print("bedbug: output_video created:", output_video)
            logger.info(f"Output video created: {output_video}")
            serializer = DoctorOutputVideoSerializer(output_video)
            # print("bedbug: Serializer data:", serializer.data)
            logger.info(f"Serializer data: {serializer.data}")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            # print(f"bedbug: Video generation failed: {e}")
            logger.error(f"Video generation failed: {e}")
            return Response({"error": "Video generation failed.", "details": str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def generate_custom_video(self, main_video_path, image_path, name, clinic, city, specialization_key, state, output_path,
                      time_duration="5-10,45-50", resolution="415x410", base_x="(main_w/2)-160", base_y="(main_h/2)-60", line_spacing="60", overlay_x="350",
                      overlay_y="70"):
        logger.info(f"Starting video generation for doctor: {name}")

        # Resource validation  
        if not os.path.exists(main_video_path):
            logger.error(f"Template video not found: {main_video_path}")
            raise Exception(f"Template video not found: {main_video_path}")

        if not os.path.exists(image_path):
            logger.error(f"Doctor image not found: {image_path}")
            raise Exception(f"Doctor image not found: {image_path}")
        # print("bedbug: Entered generate_custom_video")
        logger.info(f"Starting video generation for doctor: {name}")
        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        # print("bedbug: temp_dir created:", temp_dir)
        logger.info(f"Temp directory created: {temp_dir}")
        fps = 30
        fade_duration = 3
        # print("bedbug: Parsing time_duration for slots...")
        logger.info(f"Parsing time duration: {time_duration}")
        slots = self.parse_time_duration(time_duration)
        # print("bedbug: slots:", slots)
        logger.info(f"Parsed slots: {slots}")
        temp_videos = []
        for i, (start, end) in enumerate(slots):
            # print(f"bedbug: Creating temp video for slot {i}: start={start}, end={end}")
            logger.info(f"Creating temp video {i} for slot: {start}-{end}s")
            duration = end - start
            total_frames = duration * fps
            # print(f"bedbug: duration={duration}, total_frames={total_frames}")
            logger.debug(f"Slot {i} duration: {duration}s, frames: {total_frames}")
            zoom_effect = f"zoompan=z='1+0.00003*in':x='(iw/2)-(iw/zoom/2)':y='(ih/2)-(ih/zoom/2)':d={total_frames}:s={resolution}:fps={fps}"
            # fade_effect = f"format=rgba,fade=t=in:st=0:d={fade_duration}:alpha=1,fade=t=out:st={duration-fade_duration}:d={fade_duration}:alpha=1"
            # byvaso
            fade_out_start = max(0, duration - fade_duration)
            fade_effect = (
                f"format=rgba,"
                f"fade=t=in:st=0:d={fade_duration}:alpha=1,"
                f"fade=t=out:st={fade_out_start}:d={fade_duration}:alpha=1"
            )

            vf = f"scale={resolution},{zoom_effect},{fade_effect}"
            temp_video = os.path.join(temp_dir, f"temp_image_vid_{i}.mp4")
            # print("bedbug: Running ffmpeg for temp video:", temp_video)
            logger.debug(f"Running FFmpeg for temp video: {temp_video}")
            try:
                result = subprocess.run([
                    "ffmpeg", "-loop", "1", "-i", image_path,
                    "-vf", vf, "-t", str(duration), "-y", temp_video
                ], check=True, capture_output=True, text=True)
            except subprocess.CalledProcessError as e:
                logger.error(f"Failed to create temp video {i}: {e.stderr}")
                raise Exception(f"Failed to create temp video: {e.stderr}")
            temp_videos.append((temp_video, start, end))
        # print("bedbug: temp_videos created:", temp_videos)
        logger.info(f"Created {len(temp_videos)} temp videos")
        text_lines = [name, specialization_key, clinic, city, state]
        # font = "RobotoSlab-Medium.ttf"
# Try to find font file in different locations
        font_paths = [
            os.path.join(settings.BASE_DIR, "fonts", "RobotoSlab-Medium.ttf"),
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Common Linux font
            "/System/Library/Fonts/Arial.ttf",  # macOS fallback
            "arial.ttf"  # Windows fallback
        ]

        font = None
        for font_path in font_paths:
            if os.path.exists(font_path):
                font = font_path
                break

        if not font:
            logger.warning("No font file found, using system default")
            # Use system default font or remove font specification
            font = "arial"  # Let ffmpeg find system font
        # print("bedbug: text_lines:", text_lines)
        logger.info(f"Text lines: {text_lines}")
        # print(f"bedbug: Looking for font at: {font}")
        font_status = os.path.exists(font) if '/' in font else 'system font'
        logger.info(f"Using font: {font} (exists: {font_status})")

        # print(f"bedbug: Font exists: {os.path.exists(font) if '/' in font else 'system font'}")
        logger.info(f"Template video: {main_video_path} (exists: {os.path.exists(main_video_path)})")

        # print(f"bedbug: Template video path: {main_video_path}")
        print(f"bedbug: Template video exists: {os.path.exists(main_video_path)}")
        text_filters = []
        for start, end in slots:
            alpha_expr = f"if(lt(t\\,{start}+3),(t-{start})/3,if(lt(t\\,{end}-3),1,({end}-t)/3))"
            for j, text in enumerate(text_lines):
                offset = 132
                y_pos = f"(main_h - ({len(text_lines)}*{line_spacing}) + {j}*{line_spacing} - {offset})"
                x_pos = "(main_w/2)-300"
                # drawtext = (
                #     f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                #     f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                # )
                if font:
                    drawtext = (
                        f"drawtext=text='{text}':fontfile='{font}':fontcolor=black:fontsize=40:"
                        f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                    )
                else:
                    drawtext = (
                        f"drawtext=text='{text}':fontcolor=black:fontsize=40:"
                        f"x={x_pos}:y={y_pos}:enable='between(t,{start},{end})':alpha='{alpha_expr}'"
                    )
                print(f"bedbug: drawtext filter for line {j}:", drawtext)
                text_filters.append(drawtext)
        overlay_x1 = f"(main_w-overlay_w)/2-{int(overlay_x)}"
        overlay_y1 = f"(main_h-overlay_h)/2+{int(overlay_y)}"
        print("bedbug: overlay_x1:", overlay_x1, "overlay_y1:", overlay_y1)


        #! tODAYS WORKS AFTER DEPLOYEMNT ISSUES 
        # filter_complex = (
        #     f"[0:v][1:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[0][0]},{slots[0][1]})'[v1];"
        #     f"[v1][2:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{slots[1][0]},{slots[1][1]})'[v2];"
        #     f"[v2]{','.join(text_filters)}[v]"
        # )

        # Build overlay filters dynamically based on number of slots
        overlay_filters = []
        for i, (start, end) in enumerate(slots):
            if i == 0:
                input_label = "[0:v]"
                output_label = f"[v{i+1}]"
            else:
                input_label = f"[v{i}]"
                output_label = f"[v{i+1}]"
            
            overlay_filter = f"{input_label}[{i+1}:v]overlay=x={overlay_x1}:y={overlay_y1}:enable='between(t,{start},{end})'{output_label}"
            overlay_filters.append(overlay_filter)

        # Final filter combines overlays with text
        final_input = f"[v{len(slots)}]"
        filter_complex = f"{';'.join(overlay_filters)};{final_input}{','.join(text_filters)}[v]"
        print("bedbug: filter_complex:", filter_complex)
        # cmd = [
        #     "ffmpeg", "-i", main_video_path,
        #     "-i", temp_videos[0][0],
        #     "-i", temp_videos[1][0],
        #     "-filter_complex", filter_complex,
        #     "-map", "[v]",
        #     "-map", "0:a?",
        #     "-c:v", "libx264", "-c:a", "copy", "-y", output_path
        # ]
        cmd = ["ffmpeg", "-i", main_video_path]
        # Add all temp video inputs dynamically
        for temp_video, _, _ in temp_videos:
            cmd.extend(["-i", temp_video])

        cmd.extend([
            "-filter_complex", filter_complex,
            "-map", "[v]",
            "-map", "0:a?",
            "-c:v", "libx264", "-c:a", "copy", "-y", output_path
        ])
        print("bedbug: Running final ffmpeg command:", cmd)
        # subprocess.run(cmd, check=True)
        try:
            result = subprocess.run(cmd, check=True, capture_output=True, text=True)
        except subprocess.CalledProcessError as e:
            logger.error(f"Final FFmpeg failed: {e.stderr}")
            raise Exception(f"Video generation failed: {e.stderr}")
        for temp_video, _, _ in temp_videos:
            print("bedbug: Removing temp video:", temp_video)
            os.remove(temp_video)

    def get(self, request):
        print("bedbug: GET called, request.query_params:", request.query_params)
        doctor_id = request.query_params.get("doctor_id")
        employee_id = request.query_params.get("employee_id")
        print("bedbug: doctor_id:", doctor_id, "employee_id:", employee_id)
        videos = DoctorOutputVideo.objects.all().order_by('-id')
        if doctor_id:
            print("bedbug: Filtering by doctor_id")
            videos = videos.filter(doctor_id=doctor_id)
        if employee_id:
            print("bedbug: Filtering by employee_id")
            videos = videos.filter(doctor__employee_id=employee_id)
        paginator = Pagination_class()
        print("bedbug: Paginating videos...")
        paginated_videos = paginator.paginate_queryset(videos, request)
        serializer = DoctorOutputVideoSerializer(paginated_videos, many=True)
        print("bedbug: Returning paginated response")
        return paginator.get_paginated_response(serializer.data)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def update_employees_from_excel(request):
    excel_file = request.FILES.get('file')
    
    if not excel_file:
        return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        df = pd.read_excel(excel_file)

        updated, not_found = 0, 0
        for _, row in df.iterrows():
            employee_id = str(row['id']).strip()
            department = str(row['department']).strip()
            city = str(row['city']).strip().title()

            try:
                employee = Employee.objects.get(employee_id=employee_id)
                employee.department = department
                employee.city = city
                employee.save()
                updated += 1
            except Employee.DoesNotExist:
                not_found += 1

        return Response({
            'status': 'success',
            'updated': updated,
            'not_found': not_found
        })

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class TemplateWiseVideoCountView(APIView):
    def get(self, request):
        template_counts = DoctorOutputVideo.objects.values(
            "template__id", "template__name"
        ).annotate(
            video_count=Count("id")
        ).order_by("-video_count")

        data = [
            {
                "template_id": item["template__id"],
                "template_name": item["template__name"],
                "video_count": item["video_count"]
            }
            for item in template_counts if item["template__id"] is not None
        ]

        return Response(data, status=status.HTTP_200_OK)


